import os
import time
import random
import urllib.parse
import smtplib
from email.mime.text import MIMEText

print("✅ Standard library imports successful", flush=True)

try:
    import feedparser
    print("✅ feedparser imported", flush=True)
except ImportError as e:
    print(f"❌ Failed to import feedparser: {e}", flush=True)
    raise

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    print("✅ openpyxl imported", flush=True)
except ImportError as e:
    print(f"❌ Failed to import openpyxl: {e}", flush=True)
    raise

try:
    import requests
    print("✅ requests imported", flush=True)
except ImportError as e:
    print(f"❌ Failed to import requests: {e}", flush=True)
    raise

try:
    from google import genai
    print("✅ google-genai imported", flush=True)
except ImportError as e:
    print(f"❌ Failed to import google-genai: {e}", flush=True)
    raise

print("✅ All imports successful", flush=True)

# ==========================================
# CONFIGURATION & ENVIRONMENT SECRETS
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_KEYWORDS_FILE = os.path.join(BASE_DIR, "keywords.txt")
OUTPUT_EXCEL_FILE = os.path.join(BASE_DIR, "docs", "data", "rxbenefits_intel_hub_report.xlsx")
OUTPUT_JSON_FILE = os.path.join(BASE_DIR, "docs", "data", f"rxbenefits_{time.strftime('%Y-%m-%d')}.json")
GEMINI_API_KEY     = os.environ.get("GEMINI_API_KEY", "").strip()
WP_SITE_URL        = os.environ.get("WP_SITE_URL", "").strip()
WP_USERNAME        = os.environ.get("WP_USERNAME", "").strip()
WP_APP_PASS        = os.environ.get("WP_APP_PASS", "").strip()
SENDER_EMAIL       = os.environ.get("SENDER_EMAIL", "").strip()
SENDER_APP_PASSWORD= os.environ.get("SENDER_APP_PASSWORD", "").strip()
RECIPIENT_EMAIL    = os.environ.get("RECIPIENT_EMAIL", "").strip()

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36"
}

SYSTEM_PROMPT = """
You are an editor creating news updates for the RxBenefits Intel Hub News Radar Report. For each news article provided, follow these instructions exactly:

1. Filter the News: Only process articles that meet all of the following criteria:
   - Relevant to RxBenefits (pharmacy benefits management, self-funded employers, employee benefits, healthcare cost management, drug pricing, PBMs, specialty drugs).
   - Published by a US-based source.
   - Exclude any articles related to: Medicare, Medicaid, CMS.
   - Exclude duplicate articles or articles covering the same event/story.
2. Only give news which is highly relevant to RxBenefits.
3. Generate a New Title: Create a new, detailed, and impactful title in Title Case.
4. Write the Summary: Write one concise paragraph (maximum 5 lines) explaining What happened, Why it matters, and Key business implications.
5. Output Format:
   If the article DOES NOT pass the criteria, respond ONLY with: SKIP

   If it PASSES, return the output strictly formatted with three labeled tags:
   TITLE: <Title in Title Case>
   SUMMARY: <One-paragraph summary>
   SOURCE_LINE: Source: <Source Name>; <Publication Date>
"""


def validate_env_vars():
    """Check env vars are set — print ONLY present/missing status, never the values."""
    print("\n🔐 Validating environment variables...", flush=True)
    required = [
        "GEMINI_API_KEY",
        "WP_SITE_URL",
        "WP_USERNAME",
        "WP_APP_PASS",
        "SENDER_EMAIL",
        "SENDER_APP_PASSWORD",
        "RECIPIENT_EMAIL",
    ]
    values = {
        "GEMINI_API_KEY":      GEMINI_API_KEY,
        "WP_SITE_URL":         WP_SITE_URL,
        "WP_USERNAME":         WP_USERNAME,
        "WP_APP_PASS":         WP_APP_PASS,
        "SENDER_EMAIL":        SENDER_EMAIL,
        "SENDER_APP_PASSWORD": SENDER_APP_PASSWORD,
        "RECIPIENT_EMAIL":     RECIPIENT_EMAIL,
    }
    missing = []
    for name in required:
        if values[name]:
            # Print ONLY the length — never any characters of the value
            print(f"   ✅ {name} is set (length={len(values[name])})", flush=True)
        else:
            print(f"   ❌ {name} is NOT set or empty", flush=True)
            missing.append(name)

    if missing:
        print(f"\n❌ FATAL: Missing required env vars: {missing}", flush=True)
        return False

    print("✅ All environment variables present.\n", flush=True)
    return True


def load_keywords(filepath):
    print(f"\n📂 Looking for keywords file at: {filepath}", flush=True)

    if not os.path.exists(filepath):
        print(f"❌ ERROR: Cannot find '{filepath}'!", flush=True)
        print(f"   Files in BASE_DIR ({BASE_DIR}):", flush=True)
        for f in os.listdir(BASE_DIR):
            print(f"     - {f}", flush=True)
        return []

    with open(filepath, "r", encoding="utf-8") as f:
        keywords = [line.strip() for line in f if line.strip()]

    print(f"✅ Loaded {len(keywords)} keywords:", flush=True)
    for i, kw in enumerate(keywords, 1):
        print(f"   {i}. {kw}", flush=True)

    return keywords


def fetch_all_news(keywords):
    print(f"\n📡 Fetching news for {len(keywords)} keywords...", flush=True)
    all_articles = []

    for i, kw in enumerate(keywords, 1):
        print(f"\n   [{i}/{len(keywords)}] Fetching: '{kw}'", flush=True)
        query = f'"{kw}" when:24h'
        encoded_query = urllib.parse.quote(query)
        rss_url = f"https://news.google.com/rss/search?q={encoded_query}&hl=en-US&gl=US&ceid=US:en"

        try:
            feed = feedparser.parse(rss_url, request_headers=HEADERS)
            count = len(feed.entries)
            print(f"      Found {count} articles", flush=True)

            for item in feed.entries:
                source_info = item.get("source", {})
                source_name = (
                    source_info.get("title", "N/A")
                    if isinstance(source_info, dict)
                    else getattr(source_info, "title", "N/A")
                )
                all_articles.append({
                    "keyword":     kw,
                    "title":       item.get("title", "N/A"),
                    "link":        item.get("link", "N/A"),
                    "published":   item.get("published", item.get("pubDate", "N/A")),
                    "source_name": source_name,
                    "description": item.get("summary", item.get("description", "N/A")),
                })

        except Exception as e:
            print(f"      ❌ Error fetching '{kw}': {e}", flush=True)

        delay = random.uniform(2.0, 4.0)
        print(f"      ⏳ Sleeping {delay:.1f}s...", flush=True)
        time.sleep(delay)

    print(f"\n✅ Total raw articles fetched: {len(all_articles)}", flush=True)
    return all_articles


def process_article_with_ai(article, ai_client):
    user_prompt = f"""
    Evaluate and reformat this news article:
    - Original Title: {article['title']}
    - Source Name: {article['source_name']}
    - Publication Date: {article['published']}
    - Link: {article['link']}
    - Snippet: {article['description']}
    """
    try:
        response = ai_client.models.generate_content(
            model="gemini-flash-latest",
            contents=f"{SYSTEM_PROMPT}\n\n{user_prompt}"
        )
        text = response.text.strip()

        if text.upper().startswith("SKIP"):
            return None

        parsed = {
            "title":       "",
            "summary":     "",
            "source_line": "",
            "link":        article["link"],
            "source_name": article["source_name"],
            "date":        article["published"],
        }

        for line in text.split("\n"):
            line = line.strip()
            if line.startswith("TITLE:"):
                parsed["title"] = line.replace("TITLE:", "").strip()
            elif line.startswith("SUMMARY:"):
                parsed["summary"] = line.replace("SUMMARY:", "").strip()
            elif line.startswith("SOURCE_LINE:"):
                parsed["source_line"] = line.replace("SOURCE_LINE:", "").strip()

        # If any field is empty, skip the article
        if not parsed["title"] or not parsed["summary"] or not parsed["source_line"]:
            print(f"      ⚠️  Incomplete AI output — skipping", flush=True)
            print(f"         title={bool(parsed['title'])} summary={bool(parsed['summary'])} source={bool(parsed['source_line'])}", flush=True)
            return None

        return parsed

    except Exception as e:
        print(f"      ❌ AI Processing Error: {e}", flush=True)
        return None


def save_excel_format(processed_articles):
    print(f"\n💾 Saving {len(processed_articles)} articles to {OUTPUT_EXCEL_FILE}", flush=True)
    try:
        os.makedirs(os.path.dirname(OUTPUT_EXCEL_FILE), exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "News Radar Report"
        ws.append(["New Title", "Summary", "Source Name", "Publication Date", "Hyperlink URL"])
        for art in processed_articles:
            ws.append([art["title"], art["summary"], art["source_name"], art["date"], art["link"]])
        wb.save(OUTPUT_EXCEL_FILE)
        print("✅ Excel file saved successfully.", flush=True)
    except Exception as e:
        print(f"❌ Error saving Excel: {e}", flush=True)
        raise

def save_json_format(processed_articles):
    import json
    print(f"\n💾 Saving JSON to {OUTPUT_JSON_FILE}", flush=True)
    try:
        os.makedirs(os.path.dirname(OUTPUT_JSON_FILE), exist_ok=True)
        with open(OUTPUT_JSON_FILE, "w", encoding="utf-8") as f:
            json.dump(processed_articles, f, ensure_ascii=False, indent=2)
        print("✅ JSON file saved successfully.", flush=True)
    except Exception as e:
        print(f"❌ Error saving JSON: {e}", flush=True)
        raise
        
def upload_excel_to_wordpress(filepath):
    print(f"\n☁️  Uploading Excel to WordPress...", flush=True)
    try:
        url = f"{WP_SITE_URL}/wp-json/wp/v2/media"
        filename = os.path.basename(filepath)
        headers = {"Content-Disposition": f"attachment; filename={filename}"}

        with open(filepath, "rb") as file_data:
            response = requests.post(
                url,
                auth=(WP_USERNAME, WP_APP_PASS),
                headers=headers,
                files={"file": (
                    filename,
                    file_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )}
            )

        if response.status_code == 201:
            download_url = response.json().get("source_url")
            print(f"✅ WordPress upload successful.", flush=True)
            return download_url
        else:
            print(f"❌ WordPress upload failed. Status: {response.status_code}", flush=True)
            print(f"   Response: {response.text[:300]}", flush=True)
            return None

    except Exception as e:
        print(f"❌ WordPress upload error: {e}", flush=True)
        return None


def send_email_report(processed_articles, excel_download_url):
    from email.mime.multipart import MIMEMultipart
    from datetime import datetime

    print(f"\n📧 Sending email report ({len(processed_articles)} articles)...", flush=True)
    try:
        today = datetime.utcnow()
        subject_date = today.strftime("%m/%d/%Y")        # 07/25/2026
        body_date    = today.strftime("%B %d, %Y")       # July 25, 2026

        # ---- HTML body ----
        html_parts = []
        html_parts.append("""
<html><body style="font-family: Arial, sans-serif; font-size:14px; 
                   color:#222; max-width:680px; margin:0 auto; padding:24px;">
""")

        # Greeting
        html_parts.append(f"""
        <p style="margin: 16px 0;">Hi Lee Ashford,</p>
        <p style="margin: 0 0 20px 0;">
          Please find updates from <strong>{body_date}</strong> below:
        </p>
        <hr style="border:none; border-top:1px solid #ddd; margin-bottom:20px;">
        """)

        # Articles
        for i, art in enumerate(processed_articles, 1):
            html_parts.append(f"""
            <div style="margin-bottom:24px;">
              <div style="font-size:13px; color:#1F4E79; font-weight:500; margin-bottom:4px;">[{i}]</div>
              <div style="font-size:15px; font-weight:500; color:#000000; margin-bottom:8px; line-height:1.4;">
                {art['title']}
              </div>
              <div style="font-size:13px; color:#222; line-height:1.6; margin-bottom:8px;">
                {art['summary']}
              </div>
              <div style="font-size:12px; color:#555;">
                {art['source_line']} &nbsp;|&nbsp;
                <a href="{art['link']}" style="color:#1F4E79;">Read Full Article</a>
              </div>
            </div>
            <hr style="border:none; border-top:1px solid #eee; margin-bottom:20px;">
            """)

        # Sign-off
        html_parts.append("""
        <p style="margin-top:24px;">Regards,</p>
        <p style="font-weight:500; margin:0;">Evalueserve Team</p>

        <hr style="border:none; border-top:1px solid #ddd; margin-top:24px;">
        <div style="font-size:11px; color:#aaa; text-align:center;">
          RxBenefits Intel Hub · Daily News Radar · Automated Report
        </div>

        </body></html>
        """)

        html_body = "".join(html_parts)

        # ---- Plain text fallback ----
        text_parts = [
            f"Hi Lee Ashford,",
            f"Please find updates from {body_date} below:",
            "",
        ]
        for i, art in enumerate(processed_articles, 1):
            text_parts.append(f"[{i}] {art['title']}")
            text_parts.append(art['summary'])
            text_parts.append(f"{art['source_line']} | Link: {art['link']}")
            text_parts.append("-" * 60)
            text_parts.append("")
        text_parts += ["Regards,", "Evalueserve Team"]
        text_body = "\n".join(text_parts)

        # ---- Build message ----
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Daily News Alerts_{subject_date}"
        msg["From"]    = SENDER_EMAIL
        msg["To"]      = RECIPIENT_EMAIL

        msg.attach(MIMEText(text_body, "plain", "utf-8"))
        msg.attach(MIMEText(html_body, "html",  "utf-8"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, SENDER_APP_PASSWORD)
            server.sendmail(SENDER_EMAIL, RECIPIENT_EMAIL, msg.as_string())

        print(f"✅ Email sent. Subject: Daily News Alerts_{subject_date}", flush=True)

    except Exception as e:
        print(f"❌ Email sending error: {e}", flush=True)
        raise


# ==========================================
# MAIN
# ==========================================
if __name__ == "__main__":
    print("=" * 60, flush=True)
    print("🚀 RxBenefits Intel Hub News Radar — Script Starting", flush=True)
    print("=" * 60, flush=True)

    try:
        # Step 1: Validate env vars (prints only length, never values)
        if not validate_env_vars():
            print("❌ Exiting due to missing environment variables.", flush=True)
            exit(1)

        # Step 2: Initialize AI client (NOT at module level — would crash silently)
        print("🤖 Initializing Gemini AI client...", flush=True)
        AI_CLIENT = genai.Client(api_key=GEMINI_API_KEY)
        print("✅ Gemini AI client initialized.", flush=True)

        # Step 3: Load keywords
        keywords = load_keywords(INPUT_KEYWORDS_FILE)
        if not keywords:
            print("❌ No keywords loaded. Exiting.", flush=True)
            exit(1)

        # Step 4: Fetch news
        raw_articles = fetch_all_news(keywords)
        if not raw_articles:
            print("⚠️  No articles fetched. Exiting.", flush=True)
            exit(0)

        # Step 5: Process with AI
        print(f"\n🤖 Processing {len(raw_articles)} articles with Gemini AI...", flush=True)
        processed_articles = []
        for i, art in enumerate(raw_articles, 1):
            title_preview = art["title"][:70]
            print(f"   [{i}/{len(raw_articles)}] {title_preview}", flush=True)
            res = process_article_with_ai(art, AI_CLIENT)
            if res:
                processed_articles.append(res)
                print(f"      ✅ Kept", flush=True)
            else:
                print(f"      ⏭️  Skipped", flush=True)
            time.sleep(random.uniform(0.5, 1.5))

        print(f"\n📊 {len(processed_articles)}/{len(raw_articles)} articles kept after AI filter.", flush=True)

        if not processed_articles:
            print("⚠️  No articles passed AI filter. Exiting.", flush=True)
            exit(0)

        # Step 6: Save Excel
        save_excel_format(processed_articles)

        # Step 7: Upload to WordPress
        excel_url = upload_excel_to_wordpress(OUTPUT_EXCEL_FILE)

        # Step 8: Send email
        send_email_report(processed_articles, excel_url)

        print("\n" + "=" * 60, flush=True)
        print("✅ Process completed successfully!", flush=True)
        print("=" * 60, flush=True)

    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}", flush=True)
        import traceback
        traceback.print_exc()
        exit(1)
